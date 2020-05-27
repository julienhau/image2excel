from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path

def rgb2hex(r, g, b):
    return '{:02x}{:02x}{:02x}'.format(r, g, b)

#place image path here
p=Path('')

im = Image.open(p)
wb = openpyxl.Workbook()
sheet = wb.active
a = round(im.size[0]/150)
for x in range(0,int(im.size[0]/a)):
	for y in range(0,int(im.size[1]/a)):
		r, g, b = im.getpixel((x*a,y*a))
		hexr = rgb2hex(r, 0, 0)
		hexg = rgb2hex(0, g, 0)
		hexb = rgb2hex(0, 0, b)
		sheet[get_column_letter(x+1)+str(y*3+1)].fill=PatternFill(fgColor=hexr, fill_type='solid')
		sheet[get_column_letter(x+1)+str(y*3+2)].fill=PatternFill(fgColor=hexg, fill_type='solid')
		sheet[get_column_letter(x+1)+str(y*3+3)].fill=PatternFill(fgColor=hexb, fill_type='solid')

	print((x/im.size[0])*a*100,'%')

wb.save(p.stem+'.xlsx')
